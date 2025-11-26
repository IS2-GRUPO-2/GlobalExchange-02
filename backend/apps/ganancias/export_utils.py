"""
Utilidades para exportación de reportes de ganancias a Excel y PDF.

Genera archivos descargables con los datos de los diferentes reportes.
"""
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


def format_currency(value):
    """Formatea un valor decimal como moneda PYG."""
    if value is None:
        return "0"
    return f"Gs. {Decimal(value):,.0f}".replace(",", ".")


def format_number(value, decimals=2):
    """Formatea un número con decimales."""
    if value is None:
        return "0"
    format_str = f"{{:,.{decimals}f}}"
    return format_str.format(Decimal(value)).replace(",", ".")


def format_filtros(filtros):
    """
    Formatea los filtros aplicados para mostrar en los reportes.
    
    Args:
        filtros: dict con los filtros aplicados
        
    Returns:
        str con los filtros formateados para mostrar
    """
    if not filtros:
        return "Sin filtros aplicados"
    
    filtros_texto = []
    
    if filtros.get('fecha_inicio') and filtros.get('fecha_fin'):
        filtros_texto.append(f"Periodo: {filtros['fecha_inicio']} al {filtros['fecha_fin']}")
    elif filtros.get('anio'):
        if filtros.get('mes'):
            filtros_texto.append(f"Periodo: {filtros['mes']}/{filtros['anio']}")
        else:
            filtros_texto.append(f"Año: {filtros['anio']}")
    
    if filtros.get('divisa_nombre'):
        filtros_texto.append(f"Divisa: {filtros['divisa_nombre']}")
    
    if filtros.get('operacion'):
        operacion_texto = 'Compra' if filtros['operacion'].lower() == 'compra' else 'Venta'
        filtros_texto.append(f"Operación: {operacion_texto}")
    
    if filtros.get('metodo_nombre'):
        filtros_texto.append(f"Método: {filtros['metodo_nombre']}")
    
    if filtros.get('granularidad'):
        granularidad_texto = 'Diaria' if filtros['granularidad'] == 'dia' else 'Mensual'
        filtros_texto.append(f"Granularidad: {granularidad_texto}")
    
    return " | ".join(filtros_texto) if filtros_texto else "Sin filtros aplicados"


# ==================== EXCEL EXPORTS ====================

def export_comparativa_to_excel(data):
    """
    Exporta el reporte de comparativa de operaciones a Excel.
    
    Args:
        data: dict con estructura de ComparativaOperaciones
        
    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa Operaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "REPORTE DE COMPARATIVA DE OPERACIONES"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Headers
    headers = ['Tipo Operación', 'Total Ganancia', 'Cantidad Operaciones', 'Ganancia Promedio', 'Porcentaje del Total']
    ws.append(headers)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de compra
    compra = data.get('compra', {})
    ws.append([
        'COMPRA',
        format_currency(compra.get('total_ganancia', 0)),
        compra.get('cantidad_operaciones', 0),
        format_currency(compra.get('ganancia_promedio', 0)),
        f"{compra.get('porcentaje_total', 0):.2f}%"
    ])
    
    # Datos de venta
    venta = data.get('venta', {})
    ws.append([
        'VENTA',
        format_currency(venta.get('total_ganancia', 0)),
        venta.get('cantidad_operaciones', 0),
        format_currency(venta.get('ganancia_promedio', 0)),
        f"{venta.get('porcentaje_total', 0):.2f}%"
    ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_por_divisa_to_excel(data):
    """
    Exporta el reporte de ganancias por divisa a Excel.
    
    Args:
        data: lista de GananciaPorDivisa
        
    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ganancias por Divisa"
    
    # Estilos
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "REPORTE DE GANANCIAS POR DIVISA"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Headers
    headers = ['Código Divisa', 'Nombre Divisa', 'Total Ganancia', 'Cantidad Operaciones', 'Ganancia Promedio', 'Monto Total Operado']
    ws.append(headers)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for item in data:
        ws.append([
            item['divisa_codigo'],
            item['divisa_nombre'],
            format_currency(item['total_ganancia']),
            item['cantidad_operaciones'],
            format_currency(item['ganancia_promedio']),
            format_number(item['monto_total_operado'], 2)
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_evolucion_to_excel(data):
    """
    Exporta el reporte de evolución temporal a Excel con gráfico.
    
    Args:
        data: lista de GananciaEvolucionTemporal
        
    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Evolución Temporal"
    
    # Estilos
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "REPORTE DE EVOLUCIÓN TEMPORAL DE GANANCIAS"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Headers
    headers = ['Periodo', 'Año', 'Mes', 'Total Ganancia', 'Cantidad Operaciones', 'Ganancia Promedio']
    ws.append(headers)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for item in data:
        ws.append([
            item['periodo'],
            item['anio'],
            item['mes'],
            format_currency(item['total_ganancia']),
            item['cantidad_operaciones'],
            format_currency(item['ganancia_promedio'])
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_transacciones_to_excel(data):
    """
    Exporta el listado de transacciones a Excel.
    
    Args:
        data: lista de transacciones con sus ganancias
        
    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    # Estilos
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = "LISTADO DE TRANSACCIONES DEL PERIODO"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Headers
    headers = ['ID Transacción', 'Fecha', 'Cliente', 'Divisa', 'Operación', 
               'Monto Divisa', 'Tasa Aplicada', 'Método de Pago', 'Ganancia Neta']
    ws.append(headers)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for item in data:
        ws.append([
            item['transaccion_id'],
            item['fecha'],
            item['cliente_nombre'],
            item['divisa_codigo'],
            item['operacion'].upper(),
            format_number(item['monto_divisa'], 2),
            format_number(item['tasa_aplicada'], 2),
            item['metodo_nombre'] or 'N/A',
            format_currency(item['ganancia_neta'])
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ==================== PDF EXPORTS ====================

def export_comparativa_to_pdf(data, filtros=None):
    """
    Exporta el reporte de comparativa de operaciones a PDF.
    
    Args:
        data: dict con estructura de ComparativaOperaciones
        filtros: dict con los filtros aplicados (opcional)
        
    Returns:
        BytesIO con el archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        title='Reporte de Comparativa de Operaciones',
        author='Global Exchange'
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5496'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("REPORTE DE COMPARATIVA DE OPERACIONES", title_style)
    elements.append(title)
    
    # Fecha
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], alignment=TA_CENTER)
    date_text = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    
    # Filtros aplicados
    if filtros:
        filtros_text = format_filtros(filtros)
        filtros_style = ParagraphStyle('FiltrosStyle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor('#666666'))
        filtros_para = Paragraph(f"<b>Filtros:</b> {filtros_text}", filtros_style)
        elements.append(Spacer(1, 10))
        elements.append(filtros_para)
    
    elements.append(Spacer(1, 20))
    
    # Tabla
    compra = data.get('compra', {})
    venta = data.get('venta', {})
    
    table_data = [
        ['Tipo Operación', 'Total Ganancia', 'Cantidad\nOperaciones', 'Ganancia\nPromedio', 'Porcentaje\ndel Total'],
        [
            'COMPRA',
            format_currency(compra.get('total_ganancia', 0)),
            str(compra.get('cantidad_operaciones', 0)),
            format_currency(compra.get('ganancia_promedio', 0)),
            f"{compra.get('porcentaje_total', 0):.2f}%"
        ],
        [
            'VENTA',
            format_currency(venta.get('total_ganancia', 0)),
            str(venta.get('cantidad_operaciones', 0)),
            format_currency(venta.get('ganancia_promedio', 0)),
            f"{venta.get('porcentaje_total', 0):.2f}%"
        ]
    ]
    
    table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 1.5*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_por_divisa_to_pdf(data, filtros=None):
    """
    Exporta el reporte de ganancias por divisa a PDF.
    
    Args:
        data: lista de GananciaPorDivisa
        filtros: dict con los filtros aplicados (opcional)
        
    Returns:
        BytesIO con el archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title='Reporte de Ganancias por Divisa',
        author='Global Exchange'
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5496'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("REPORTE DE GANANCIAS POR DIVISA", title_style)
    elements.append(title)
    
    # Fecha
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], alignment=TA_CENTER)
    date_text = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    
    # Filtros aplicados
    if filtros:
        filtros_text = format_filtros(filtros)
        filtros_style = ParagraphStyle('FiltrosStyle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor('#666666'))
        filtros_para = Paragraph(f"<b>Filtros:</b> {filtros_text}", filtros_style)
        elements.append(Spacer(1, 10))
        elements.append(filtros_para)
    
    elements.append(Spacer(1, 20))
    
    # Tabla
    table_data = [['Código', 'Nombre Divisa', 'Total Ganancia', 'Cantidad\nOperaciones', 'Ganancia\nPromedio', 'Monto Total\nOperado']]
    
    for item in data:
        table_data.append([
            item['divisa_codigo'],
            item['divisa_nombre'],
            format_currency(item['total_ganancia']),
            str(item['cantidad_operaciones']),
            format_currency(item['ganancia_promedio']),
            format_number(item['monto_total_operado'], 2)
        ])
    
    table = Table(table_data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1.2*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_evolucion_to_pdf(data, filtros=None):
    """
    Exporta el reporte de evolución temporal a PDF.
    
    Args:
        data: lista de GananciaEvolucionTemporal
        filtros: dict con los filtros aplicados (opcional)
        
    Returns:
        BytesIO con el archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        title='Reporte de Evolución Temporal',
        author='Global Exchange'
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5496'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("REPORTE DE EVOLUCIÓN TEMPORAL", title_style)
    elements.append(title)
    
    # Fecha
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], alignment=TA_CENTER)
    date_text = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    
    # Filtros aplicados
    if filtros:
        filtros_text = format_filtros(filtros)
        filtros_style = ParagraphStyle('FiltrosStyle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor('#666666'))
        filtros_para = Paragraph(f"<b>Filtros:</b> {filtros_text}", filtros_style)
        elements.append(Spacer(1, 10))
        elements.append(filtros_para)
    
    elements.append(Spacer(1, 20))
    
    # Tabla
    table_data = [['Periodo', 'Año', 'Mes', 'Total Ganancia', 'Cantidad\nOperaciones', 'Ganancia\nPromedio']]
    
    for item in data:
        table_data.append([
            item['periodo'],
            str(item['anio']),
            str(item['mes']),
            format_currency(item['total_ganancia']),
            str(item['cantidad_operaciones']),
            format_currency(item['ganancia_promedio'])
        ])
    
    table = Table(table_data, colWidths=[1*inch, 0.8*inch, 0.8*inch, 1.5*inch, 1.2*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_transacciones_to_pdf(data, filtros=None):
    """
    Exporta el listado de transacciones a PDF.
    
    Args:
        data: lista de transacciones con sus ganancias
        filtros: dict con los filtros aplicados (opcional)
        
    Returns:
        BytesIO con el archivo PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title='Listado de Transacciones',
        author='Global Exchange'
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5496'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("LISTADO DE TRANSACCIONES DEL PERIODO", title_style)
    elements.append(title)
    
    # Fecha
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], alignment=TA_CENTER)
    date_text = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    
    # Filtros aplicados
    if filtros:
        filtros_text = format_filtros(filtros)
        filtros_style = ParagraphStyle('FiltrosStyle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor('#666666'))
        filtros_para = Paragraph(f"<b>Filtros:</b> {filtros_text}", filtros_style)
        elements.append(Spacer(1, 10))
        elements.append(filtros_para)
    
    elements.append(Spacer(1, 20))
    
    # Tabla
    table_data = [['ID', 'Fecha', 'Cliente', 'Divisa', 'Operación', 'Monto', 'Tasa', 'Método', 'Ganancia']]
    
    for item in data:
        table_data.append([
            str(item['transaccion_id']),
            item['fecha'],
            item['cliente_nombre'][:25],  # Limitar longitud
            item['divisa_codigo'],
            item['operacion'].upper(),
            format_number(item['monto_divisa'], 2),
            format_number(item['tasa_aplicada'], 2),
            (item['metodo_nombre'] or 'N/A')[:15],
            format_currency(item['ganancia_neta'])
        ])
    
    table = Table(table_data, colWidths=[0.5*inch, 0.7*inch, 1.4*inch, 0.5*inch, 0.7*inch, 
                                          0.9*inch, 0.9*inch, 1*inch, 1.1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
